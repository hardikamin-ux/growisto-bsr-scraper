import os
import sys
import io
import tempfile
import subprocess
import datetime
import threading
import time
from typing import Optional

# Global lock — only one scrape job runs at a time across all sessions
_scraper_lock = threading.Lock()
_scraper_active = threading.Event()  # set while a job is running

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Growisto BSR Scraper",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Growisto Brand CSS ────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700;800&display=swap');

html, body, [class*="css"], * { font-family: 'Poppins', sans-serif !important; }
.stApp { background: #F6F6F4 !important; }
#MainMenu, footer, header { visibility: hidden; }

/* ── Native Streamlit containers — give them a card look ── */
[data-testid="stVerticalBlockBorderWrapper"] {
  border: 1.5px solid #B8DBD9 !important;
  border-radius: 12px !important;
  background: #FFFFFF !important;
  padding: 4px !important;
  box-shadow: 0 1px 8px rgba(54,117,136,0.08) !important;
}

/* ── Header card ── */
.g-header {
  background: #367588;
  border-radius: 12px;
  padding: 20px 28px;
  margin-bottom: 4px;
  display: flex; align-items: center; gap: 16px;
}
.g-header-icon {
  background: rgba(255,255,255,0.18);
  border-radius: 10px; width: 46px; height: 46px;
  display: flex; align-items: center; justify-content: center;
  font-size: 24px;
}
.g-header-title {
  font-size: 19px; font-weight: 700; color: #FFFFFF;
  margin: 0; line-height: 1.3;
}
.g-header-sub {
  font-size: 12px; color: rgba(255,255,255,0.78);
  margin: 1px 0 0; font-weight: 400;
}
.g-badge {
  background: #E35D34; color: #fff;
  border-radius: 20px; padding: 2px 9px;
  font-size: 10px; font-weight: 700; margin-left: 8px;
  vertical-align: middle; letter-spacing: 0.3px;
}

/* ── Section labels ── */
.g-label {
  font-size: 11px; font-weight: 700; color: #367588;
  text-transform: uppercase; letter-spacing: 0.8px;
  margin: 0 0 2px; border-left: 3px solid #367588;
  padding-left: 8px;
}
.g-sublabel {
  font-size: 12px; color: #1D1D20; opacity: 0.65;
  margin: 0 0 14px;
}

/* ── Info / status banners ── */
.g-success {
  background: #B8DBD9; border-left: 4px solid #367588;
  border-radius: 8px; padding: 12px 16px; margin: 8px 0;
  color: #1D1D20 !important; font-weight: 600; font-size: 13px;
}
.g-error {
  background: #fdecea; border-left: 4px solid #E35D34;
  border-radius: 8px; padding: 12px 16px; margin: 8px 0;
  color: #a93220 !important; font-weight: 600; font-size: 13px;
}
.g-info {
  background: #eaf4f5; border-left: 4px solid #367588;
  border-radius: 8px; padding: 12px 16px; margin: 8px 0;
  color: #367588 !important; font-weight: 500; font-size: 13px;
}

/* ── Steps (how-to) ── */
.g-step { display: flex; align-items: flex-start; gap: 12px; margin-bottom: 14px; }
.g-step-num {
  background: #367588; color: #fff;
  min-width: 26px; height: 26px; border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  font-weight: 700; font-size: 12px;
}
.g-step-text { font-size: 12px; color: #1D1D20 !important; line-height: 1.55; padding-top: 3px; }
.g-step-text b { color: #367588; }

/* ── Marketplace chips ── */
.g-chip {
  display: inline-block; background: #B8DBD9; color: #1D1D20;
  border-radius: 20px; padding: 3px 11px; font-size: 11px;
  font-weight: 600; margin: 3px 2px;
}

/* ── Stat tiles ── */
.g-stats { display: flex; gap: 10px; margin: 12px 0; }
.g-stat {
  flex: 1; background: #F6F6F4; border-radius: 10px;
  padding: 14px 16px; border-top: 3px solid #367588; text-align: center;
}
.g-stat-flame { border-top-color: #E35D34; }
.g-stat-num { font-size: 26px; font-weight: 800; color: #367588; margin: 0; line-height: 1; }
.g-stat-num-flame { color: #E35D34; }
.g-stat-label { font-size: 10px; color: #1D1D20; font-weight: 600;
  margin: 4px 0 0; text-transform: uppercase; letter-spacing: 0.5px; }

/* ── Log terminal ── */
.log-box {
  background: #1D1D20; color: #B8DBD9; border-radius: 8px;
  padding: 14px 16px; font-family: 'Courier New', monospace !important;
  font-size: 11px; line-height: 1.6;
  max-height: 260px; overflow-y: auto;
  white-space: pre-wrap; border: 1px solid #367588;
}

/* ── Buttons ── */
.stButton > button {
  background: #367588 !important; color: #FFFFFF !important;
  border: none !important; border-radius: 8px !important;
  font-weight: 600 !important; font-size: 13px !important;
  font-family: 'Poppins', sans-serif !important;
  padding: 9px 22px !important; width: 100% !important;
}
.stButton > button:hover {
  background: #2a5c6a !important;
  box-shadow: 0 4px 14px rgba(54,117,136,0.3) !important;
}
.stDownloadButton > button {
  background: #E35D34 !important; color: #FFFFFF !important;
  border: none !important; border-radius: 8px !important;
  font-weight: 700 !important; font-size: 14px !important;
  font-family: 'Poppins', sans-serif !important;
  padding: 11px 22px !important; width: 100% !important;
}
.stDownloadButton > button:hover { background: #c44d27 !important; }

/* ── Inputs ── */
.stTextInput input, .stTextArea textarea {
  border: 1.5px solid #B8DBD9 !important; border-radius: 8px !important;
  color: #1D1D20 !important; background: #FFFFFF !important;
  font-size: 13px !important; font-family: 'Poppins', sans-serif !important;
}
.stTextInput input:focus, .stTextArea textarea:focus {
  border-color: #367588 !important;
  box-shadow: 0 0 0 3px rgba(54,117,136,0.12) !important;
}
input[type="password"] { color: #1D1D20 !important; background: #FFFFFF !important; }
.stSelectbox > div > div {
  border: 1.5px solid #B8DBD9 !important; border-radius: 8px !important;
  background: #FFFFFF !important; color: #1D1D20 !important;
}
::placeholder { color: #9CA3AF !important; }
label { color: #1D1D20 !important; font-weight: 500 !important; font-size: 13px !important; }

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
  background: #F6F6F4; border-radius: 10px; padding: 4px; gap: 5px;
  border: 1.5px solid #B8DBD9;
}
.stTabs [data-baseweb="tab"] {
  border-radius: 7px !important; font-weight: 500 !important;
  font-size: 13px !important; color: #1D1D20 !important; padding: 7px 18px !important;
}
.stTabs [aria-selected="true"] {
  background: #367588 !important; color: #FFFFFF !important; font-weight: 600 !important;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] > div {
  border: 2px dashed #B8DBD9 !important; border-radius: 10px !important;
  background: #F6F6F4 !important;
}

/* ── Progress bar ── */
.stProgress > div > div > div > div { background: #367588 !important; }
.stProgress > div > div > div { background: #B8DBD9 !important; border-radius: 99px !important; }

/* ── Login ── */
.g-login-top {
  background: #367588; border-radius: 14px 14px 0 0;
  padding: 32px 36px 24px; text-align: center;
}
.g-login-body { padding: 28px 36px 32px; background: #fff; border-radius: 0 0 14px 14px; }
.g-login-logo { font-size: 40px; margin-bottom: 10px; }
.g-login-brand { font-size: 20px; font-weight: 800; color: #fff; margin: 0; }
.g-login-tagline { font-size: 12px; color: rgba(255,255,255,0.75); margin: 3px 0 0; }
.g-login-title { font-size: 14px; font-weight: 600; color: #1D1D20; margin: 0 0 16px; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
PASSWORD    = "Growisto@2026"
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
SCRIPT_PATH = os.path.join(SCRIPT_DIR, "bsr_scraper.py")
MARKETPLACE_LABELS = {
    "US": "🇺🇸  United States", "UK": "🇬🇧  United Kingdom",
    "DE": "🇩🇪  Germany",       "IN": "🇮🇳  India",
    "CA": "🇨🇦  Canada",        "AU": "🇦🇺  Australia",
    "FR": "🇫🇷  France",        "ES": "🇪🇸  Spain",
    "IT": "🇮🇹  Italy",         "JP": "🇯🇵  Japan",
    "MX": "🇲🇽  Mexico",        "AE": "🇦🇪  UAE",
}

for k, v in {"authenticated": False, "output_bytes": None}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Template builder (Growisto branded) ──────────────────────────────────────

def get_blank_template_bytes() -> bytes:
    TEAL   = "FF367588"
    POWDER = "FFB8DBD9"
    RAISIN = "FF1D1D20"
    WHITE  = "FFFFFFFF"
    CULTRD = "FFF6F6F4"
    FONT   = "Poppins"

    thin   = Side(style="thin", color="B8DBD9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left_a = Alignment(horizontal="left",   vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Product List"

    # Row 1 — brand title
    ws.merge_cells("A1:D1")
    ws["A1"] = "  Growisto BSR Scraper — Product Input Template"
    ws["A1"].fill      = PatternFill("solid", fgColor=TEAL)
    ws["A1"].font      = Font(bold=True, color=WHITE, name=FONT, size=12)
    ws["A1"].alignment = left_a
    ws.row_dimensions[1].height = 30

    # Row 2 — instruction
    ws.merge_cells("A2:D2")
    ws["A2"] = "  Fill in ASIN and select Marketplace from dropdown. Valid codes: US UK DE IN CA AU FR ES IT JP MX AE"
    ws["A2"].fill      = PatternFill("solid", fgColor=POWDER)
    ws["A2"].font      = Font(italic=True, color=RAISIN, name=FONT, size=9)
    ws["A2"].alignment = left_a
    ws.row_dimensions[2].height = 16

    # Row 3 — column headers
    for col, h in enumerate(["#", "ASIN", "Marketplace Code", "Notes (optional)"], 1):
        cell           = ws.cell(row=3, column=col, value=h)
        cell.fill      = PatternFill("solid", fgColor=TEAL)
        cell.font      = Font(bold=True, color=WHITE, name=FONT, size=10)
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[3].height = 22

    # Dropdown validation
    dv = DataValidation(
        type="list",
        formula1='"US,UK,DE,IN,CA,AU,FR,ES,IT,JP,MX,AE"',
        allow_blank=True, showDropDown=False,
        showErrorMessage=True, errorTitle="Invalid Code",
        error="Pick a code from the dropdown."
    )
    ws.add_data_validation(dv)
    dv.sqref = "C4:C103"

    # 100 data rows
    for r in range(4, 104):
        row_fill = PatternFill("solid", fgColor=CULTRD if r % 2 == 0 else "FFFFFFFF")
        ws.cell(row=r, column=1, value=r - 3).fill = row_fill
        ws.cell(row=r, column=1).font      = Font(name=FONT, size=9, color=RAISIN)
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=1).border    = border
        for c in range(2, 5):
            cell           = ws.cell(row=r, column=c)
            cell.fill      = row_fill
            cell.font      = Font(name=FONT, size=9, color=RAISIN)
            cell.alignment = center
            cell.border    = border
        ws.row_dimensions[r].height = 18

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 32
    ws.freeze_panes = "A4"

    # Second sheet — marketplace reference
    ref = wb.create_sheet("Marketplace Codes")
    ref.merge_cells("A1:C1")
    ref["A1"] = "  Supported Marketplace Codes"
    ref["A1"].fill      = PatternFill("solid", fgColor=TEAL)
    ref["A1"].font      = Font(bold=True, color=WHITE, name=FONT, size=11)
    ref["A1"].alignment = left_a
    ref.row_dimensions[1].height = 26

    rows_data = [
        ("Code", "Amazon Domain", "Country"),
        ("US", "amazon.com", "United States"),
        ("UK", "amazon.co.uk", "United Kingdom"),
        ("DE", "amazon.de", "Germany"),
        ("IN", "amazon.in", "India"),
        ("CA", "amazon.ca", "Canada"),
        ("AU", "amazon.com.au", "Australia"),
        ("FR", "amazon.fr", "France"),
        ("ES", "amazon.es", "Spain"),
        ("IT", "amazon.it", "Italy"),
        ("JP", "amazon.co.jp", "Japan"),
        ("MX", "amazon.com.mx", "Mexico"),
        ("AE", "amazon.ae", "UAE"),
    ]
    for r, row in enumerate(rows_data, 2):
        for c, val in enumerate(row, 1):
            cell           = ref.cell(row=r, column=c, value=val)
            cell.border    = border
            cell.alignment = center
            if r == 2:
                cell.fill = PatternFill("solid", fgColor=TEAL)
                cell.font = Font(bold=True, color=WHITE, name=FONT, size=10)
            else:
                cell.fill = PatternFill("solid", fgColor=CULTRD if r % 2 == 0 else "FFFFFFFF")
                cell.font = Font(name=FONT, size=9, color=RAISIN)
        ref.row_dimensions[r].height = 18
    ref.column_dimensions["A"].width = 10
    ref.column_dimensions["B"].width = 22
    ref.column_dimensions["C"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Input Excel builder ───────────────────────────────────────────────────────

def make_input_excel(asins: list, marketplace: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product List"
    ws.cell(row=1, column=1, value="BSR Scraper — Product Input Template")
    ws.cell(row=2, column=1, value="Auto-generated by app")
    ws.cell(row=3, column=1, value="#")
    ws.cell(row=3, column=2, value="ASIN")
    ws.cell(row=3, column=3, value="Marketplace Code")
    ws.cell(row=3, column=4, value="Notes (optional)")
    for i, asin in enumerate(asins, 1):
        ws.cell(row=i + 3, column=1, value=i)
        ws.cell(row=i + 3, column=2, value=asin.strip().upper())
        ws.cell(row=i + 3, column=3, value=marketplace.upper())
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


# ── Scraper runner ────────────────────────────────────────────────────────────

def run_scraper(input_path: str, log_ph, prog_ph, stat_ph) -> Optional[bytes]:
    out_tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    out_tmp.close()
    cmd  = [sys.executable, SCRIPT_PATH, "--input", input_path, "--output", out_tmp.name]
    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                            text=True, bufsize=1, cwd=SCRIPT_DIR)
    log_lines, success_count, failed_count, total = [], 0, 0, 0
    for line in proc.stdout:
        clean = line.rstrip()
        if not clean:
            continue
        log_lines.append(clean)
        if "product(s) to scrape" in clean:
            try:
                total = int(clean.split("Found")[1].split("product")[0].strip())
            except Exception:
                pass
        if "  OK  " in clean:
            success_count += 1
        if "FAILED —" in clean:
            failed_count += 1
        log_ph.markdown(f'<div class="log-box">{"chr(10)".join(log_lines[-20:])}</div>',
                        unsafe_allow_html=True)
        if total > 0:
            prog_ph.progress(min((success_count + failed_count) / total, 1.0))
        stat_ph.markdown(f"""
        <div class="g-stats">
          <div class="g-stat"><p class="g-stat-num">{total}</p><p class="g-stat-label">Total</p></div>
          <div class="g-stat"><p class="g-stat-num">{success_count}</p><p class="g-stat-label">Scraped</p></div>
          <div class="g-stat g-stat-flame"><p class="g-stat-num g-stat-num-flame">{failed_count}</p><p class="g-stat-label">Failed</p></div>
        </div>""", unsafe_allow_html=True)
    proc.wait()
    if proc.returncode == 0 and os.path.exists(out_tmp.name):
        with open(out_tmp.name, "rb") as f:
            data = f.read()
        os.unlink(out_tmp.name)
        return data
    try:
        os.unlink(out_tmp.name)
    except Exception:
        pass
    return None


# ── Results table renderer ───────────────────────────────────────────────────

def render_results_table(output_bytes: bytes) -> None:
    """Parse the output Excel and render a branded HTML results table."""
    buf = io.BytesIO(output_bytes)
    wb  = load_workbook(buf)
    ws  = wb.active

    # Row 3 = headers, rows 4+ = data (skip title rows 1-2 and summary last row)
    headers = [ws.cell(row=3, column=c).value for c in range(1, 13)]
    rows = []
    for r in range(4, ws.max_row):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 13)]
        if any(v is not None for v in row_vals):
            rows.append(row_vals)

    if not rows:
        return

    # Build HTML table
    cols_to_show = [0, 1, 2, 3, 4, 5, 6, 7, 10]  # #, ASIN, Market, Title, BSR x2, Sub x2, Status
    col_labels   = ["#", "ASIN", "Market", "Product Title",
                    "Main BSR Rank", "Main BSR Category",
                    "Sub BSR Rank", "Sub BSR Category", "Status"]
    col_widths   = ["4%","10%","7%","28%","9%","16%","9%","12%","5%"]

    header_row = "".join(
        f'<th style="background:#367588;color:#fff;padding:10px 12px;font-weight:600;'
        f'font-size:12px;text-align:{"left" if i==3 else "center"};'
        f'width:{col_widths[i]};white-space:nowrap;">{col_labels[i]}</th>'
        for i, _ in enumerate(cols_to_show)
    )

    data_rows_html = ""
    for i, row in enumerate(rows):
        status     = str(row[10] or "")
        is_failed  = "Failed" in status or "FAILED" in status
        row_bg     = "#fdecea" if is_failed else ("#F6F6F4" if i % 2 == 0 else "#EAF4F5")
        status_bg  = "#E35D34" if is_failed else "#367588"
        status_txt = "✗ Failed" if is_failed else "✓ OK"

        cells = ""
        for j, col_idx in enumerate(cols_to_show):
            val = row[col_idx]
            if val is None:
                val = ""
            # Format rank numbers with comma separator
            if col_idx in (4, 6) and val and str(val).isdigit():
                val = f"#{int(val):,}"
            # Status cell — coloured badge
            if col_idx == 10:
                cells += (f'<td style="text-align:center;padding:8px 6px;">'
                          f'<span style="background:{status_bg};color:#fff;border-radius:12px;'
                          f'padding:3px 10px;font-size:11px;font-weight:600;white-space:nowrap;">'
                          f'{status_txt}</span></td>')
            else:
                align = "left" if col_idx == 3 else "center"
                bold  = "700" if col_idx in (4, 6) else "400"
                color = "#367588" if col_idx in (4, 6) else "#1D1D20"
                cells += (f'<td style="padding:8px 12px;font-size:12px;color:{color};'
                          f'font-weight:{bold};text-align:{align};'
                          f'max-width:{"240px" if col_idx==3 else "auto"};'
                          f'white-space:{"normal" if col_idx==3 else "nowrap"};">'
                          f'{val}</td>')

        data_rows_html += f'<tr style="background:{row_bg};">{cells}</tr>'

    table_html = f"""
    <div style="overflow-x:auto;border-radius:10px;border:1.5px solid #B8DBD9;margin-top:8px;">
      <table style="width:100%;border-collapse:collapse;font-family:Poppins,sans-serif;">
        <thead><tr>{header_row}</tr></thead>
        <tbody>{data_rows_html}</tbody>
      </table>
    </div>
    """
    st.markdown(table_html, unsafe_allow_html=True)


# ── Login ─────────────────────────────────────────────────────────────────────

def show_login():
    _, col, _ = st.columns([1, 1.1, 1])
    with col:
        st.markdown("""
        <div style="box-shadow:0 8px 40px rgba(54,117,136,0.18); border-radius:14px; overflow:hidden; margin-top:40px;">
          <div class="g-login-top">
            <div class="g-login-logo">📊</div>
            <p class="g-login-brand">Growisto BSR Scraper</p>
            <p class="g-login-tagline">Amazon Best Seller Rank Intelligence Tool</p>
          </div>
          <div class="g-login-body">
            <p class="g-login-title">Enter your password to continue</p>
        """, unsafe_allow_html=True)
        pwd = st.text_input("Password", type="password",
                            placeholder="Password", label_visibility="collapsed")
        if st.button("Sign In →", key="login_btn"):
            if pwd == PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.markdown('<div class="g-error">❌ Incorrect password.</div>', unsafe_allow_html=True)
        st.markdown('</div></div>', unsafe_allow_html=True)


# ── Main App ──────────────────────────────────────────────────────────────────

def show_app():

    # Header banner
    st.markdown("""
    <div class="g-header">
      <div class="g-header-icon">📊</div>
      <div>
        <p class="g-header-title">Growisto BSR Scraper
          <span class="g-badge">v1.0</span>
        </p>
        <p class="g-header-sub">Amazon Best Seller Rank Intelligence — any ASIN, any marketplace</p>
      </div>
    </div>
    """, unsafe_allow_html=True)

    left_col, right_col = st.columns([1.4, 1], gap="large")
    input_path, ready = None, False

    with left_col:
        with st.container(border=True):
            st.markdown('<p class="g-label">📥 Input</p>', unsafe_allow_html=True)
            st.markdown('<p class="g-sublabel">Enter ASINs manually or upload your template</p>', unsafe_allow_html=True)

            tab1, tab2 = st.tabs(["✏️  Paste ASINs", "📂  Upload Template"])

            with tab1:
                asin_text = st.text_area("ASINs — one per line", height=170,
                                         placeholder="B08T3MR7Z9\nB0BBPQHGR6\nB0BRYC3FDX")
                mp_label  = st.selectbox("Marketplace", list(MARKETPLACE_LABELS.values()))
                mp_code   = [k for k, v in MARKETPLACE_LABELS.items() if v == mp_label][0]
                if st.button("🚀  Run Scraper", key="run_paste"):
                    asins = [a.strip().upper() for a in asin_text.splitlines() if a.strip()]
                    if not asins:
                        st.markdown('<div class="g-error">Paste at least one ASIN.</div>', unsafe_allow_html=True)
                    else:
                        input_path = make_input_excel(asins, mp_code)
                        ready = True

            with tab2:
                st.markdown('<p style="font-size:12px;color:#1D1D20;margin-bottom:8px;">Don\'t have the template? Download it first, fill in your ASINs, then upload.</p>', unsafe_allow_html=True)
                st.download_button(
                    "⬇️  Download Blank Template",
                    data=get_blank_template_bytes(),
                    file_name="BSR_Input_Template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_tpl",
                )
                st.markdown("<br>", unsafe_allow_html=True)
                uploaded = st.file_uploader("Upload filled template (.xlsx)", type=["xlsx"])
                if uploaded:
                    tmp_up = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                    tmp_up.write(uploaded.read())
                    tmp_up.close()
                    try:
                        wb  = load_workbook(tmp_up.name)
                        ws  = wb.active
                        rows = [r for r in ws.iter_rows(min_row=4, values_only=True)
                                if r[1] and str(r[1]).strip()
                                and not (r[3] and "example" in str(r[3]).lower())]
                        if rows:
                            st.markdown(f'<div class="g-info">✅ {len(rows)} ASIN(s) detected</div>', unsafe_allow_html=True)
                            if st.button("🚀  Run Scraper", key="run_upload"):
                                input_path = tmp_up.name
                                ready = True
                        else:
                            st.markdown('<div class="g-error">No valid ASINs found in file.</div>', unsafe_allow_html=True)
                    except Exception:
                        st.markdown('<div class="g-error">Could not read file.</div>', unsafe_allow_html=True)

    with right_col:
        with st.container(border=True):
            st.markdown("""
            <p class="g-label">🗺️ How to use</p>
            <br>
            <div class="g-step"><div class="g-step-num">1</div>
              <div class="g-step-text">Paste ASINs (one per line) <b>or</b> download &amp; fill the template for mixed marketplaces</div></div>
            <div class="g-step"><div class="g-step-num">2</div>
              <div class="g-step-text">Select the <b>marketplace</b> (or set per-row in the template)</div></div>
            <div class="g-step"><div class="g-step-num">3</div>
              <div class="g-step-text">Click <b>Run Scraper</b> — Chrome opens and scrapes each page automatically</div></div>
            <div class="g-step"><div class="g-step-num">4</div>
              <div class="g-step-text"><b>Download</b> the Growisto-branded Excel report when done</div></div>
            """, unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown("""
            <p class="g-label">🌍 Supported Marketplaces</p><br>
            <span class="g-chip">🇺🇸 US</span><span class="g-chip">🇬🇧 UK</span>
            <span class="g-chip">🇩🇪 DE</span><span class="g-chip">🇮🇳 IN</span>
            <span class="g-chip">🇨🇦 CA</span><span class="g-chip">🇦🇺 AU</span>
            <span class="g-chip">🇫🇷 FR</span><span class="g-chip">🇪🇸 ES</span>
            <span class="g-chip">🇮🇹 IT</span><span class="g-chip">🇯🇵 JP</span>
            <span class="g-chip">🇲🇽 MX</span><span class="g-chip">🇦🇪 AE</span>
            """, unsafe_allow_html=True)

    # ── Results section
    if ready and input_path:
        st.divider()
        with st.container(border=True):

            # ── Wait if another job is running ────────────────────────────
            if _scraper_active.is_set():
                st.markdown('<p class="g-label">⏳ Waiting in Queue</p>', unsafe_allow_html=True)
                wait_ph = st.empty()
                wait_secs = 0
                while _scraper_active.is_set():
                    wait_ph.markdown(
                        f'<div class="g-info">🕐 Another scrape is currently running. '
                        f'You\'re next — please wait... ({wait_secs}s)</div>',
                        unsafe_allow_html=True
                    )
                    time.sleep(3)
                    wait_secs += 3
                    st.rerun()
                wait_ph.empty()

            # ── Acquire lock and run ──────────────────────────────────────
            _scraper_active.set()
            try:
                st.markdown('<p class="g-label">⚙️ Scraping in Progress</p>', unsafe_allow_html=True)
                st.markdown('<p class="g-sublabel">Running headless Chrome on the server</p>', unsafe_allow_html=True)
                prog_ph = st.progress(0)
                stat_ph = st.empty()
                log_ph  = st.empty()
                with st.spinner(""):
                    output_bytes = run_scraper(input_path, log_ph, prog_ph, stat_ph)
            finally:
                _scraper_active.clear()
                if input_path and os.path.exists(input_path):
                    try:
                        os.unlink(input_path)
                    except Exception:
                        pass

            prog_ph.progress(1.0)
            if output_bytes:
                st.markdown('<div class="g-success">✅ Scraping complete!</div>', unsafe_allow_html=True)

                # ── Inline results table
                st.markdown('<p class="g-label" style="margin-top:20px;">📋 Results</p>', unsafe_allow_html=True)
                render_results_table(output_bytes)

                # ── Download button
                st.markdown("<br>", unsafe_allow_html=True)
                _, dl_col, _ = st.columns([1, 1.4, 1])
                with dl_col:
                    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    st.download_button(
                        "⬇️  Download BSR Report (.xlsx)",
                        data=output_bytes,
                        file_name=f"Growisto_BSR_Report_{ts}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            else:
                st.markdown('<div class="g-error">❌ Scraper failed. Check the log above.</div>', unsafe_allow_html=True)


# ── Entry ─────────────────────────────────────────────────────────────────────
show_app()
