#!/usr/bin/env python3
"""
BSR Scraper — Playwright edition
Opens a real Chrome browser, navigates to each ASIN, scrolls to the product
details section, extracts Best Seller Rank data, and writes results to Excel.
"""

import os
import re
import time
import random
import datetime
from typing import List, Tuple, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_FILE  = "BSR_Input_Template.xlsx"
MAX_RETRIES = 3

DOMAINS = {
    "US": "https://www.amazon.com",
    "UK": "https://www.amazon.co.uk",
    "DE": "https://www.amazon.de",
    "IN": "https://www.amazon.in",
    "CA": "https://www.amazon.ca",
    "AU": "https://www.amazon.com.au",
    "FR": "https://www.amazon.fr",
    "ES": "https://www.amazon.es",
    "IT": "https://www.amazon.it",
    "JP": "https://www.amazon.co.jp",
    "MX": "https://www.amazon.com.mx",
    "AE": "https://www.amazon.ae",
}

# Locale + timezone per marketplace so Amazon serves the correct language
LOCALE_MAP = {
    "US": ("en-US", "America/New_York"),
    "UK": ("en-GB", "Europe/London"),
    "DE": ("de-DE", "Europe/Berlin"),
    "FR": ("fr-FR", "Europe/Paris"),
    "ES": ("es-ES", "Europe/Madrid"),
    "IT": ("it-IT", "Europe/Rome"),
    "CA": ("en-CA", "America/Toronto"),
    "AU": ("en-AU", "Australia/Sydney"),
    "JP": ("ja-JP", "Asia/Tokyo"),
    "IN": ("en-IN", "Asia/Kolkata"),
    "MX": ("es-MX", "America/Mexico_City"),
    "AE": ("ar-AE", "Asia/Dubai"),
}

# ── BSR Parsing ───────────────────────────────────────────────────────────────

# BSR section label in every supported language
BSR_LABELS = [
    "best sellers rank",                           # EN (US, UK, CA, AU, IN)
    "amazon bestseller-rang",                      # DE
    "classement des meilleures ventes",            # FR
    "clasificación en los más vendidos",           # ES
    "posizione nella classifica bestseller",       # IT
    "売れ筋ランキング",                               # JP
    "en iyi satıcılar sıralaması",                 # TR
    "en çok satanlar sıralaması",                  # TR alt
]

def is_bsr_label(text: str) -> bool:
    t = text.lower()
    return any(label in t for label in BSR_LABELS)


def extract_ranks(text: str) -> List[Tuple[str, str]]:
    """
    Handles all marketplace rank formats:
      EN:  #1,234 in Category
      DE:  Nr. 5 in Fashion
      ES:  nº3 en Moda
      IT:  n. 1 in Moda
      FR:  1 en Mode  /  N° 1 en Mode
      JP:  1位 Category
    """
    results = []

    # Pattern 1 — English:  #1,234 in Category
    for m in re.finditer(r'#([\d,]+)\s+in\s+([^#\(\n]+?)(?=\s*\(|\s*#|\s*$)', text):
        rank = m.group(1).replace(",", "")
        cat  = m.group(2).strip().rstrip("(").strip()
        if rank and cat:
            results.append((rank, cat))

    if results:
        return results

    # Pattern 2 — German:  Nr. 5 in Fashion
    for m in re.finditer(r'Nr\.\s*([\d\.]+)\s+in\s+([^\(\n]+?)(?=\s*\(|\s*Nr\.|\s*$)', text):
        rank = m.group(1).replace(".", "")
        cat  = m.group(2).strip()
        if rank and cat:
            results.append((rank, cat))

    if results:
        return results

    # Pattern 3 — Spanish:  nº3 en Moda
    for m in re.finditer(r'nº\s*([\d]+)\s+en\s+([^\(\n]+?)(?=\s*\(|\s*nº|\s*$)', text, re.IGNORECASE):
        rank = m.group(1)
        cat  = m.group(2).strip()
        if rank and cat:
            results.append((rank, cat))

    if results:
        return results

    # Pattern 4 — Italian:  n. 1 in Moda
    for m in re.finditer(r'n\.\s*([\d]+)\s+in\s+([^\(\n]+?)(?=\s*\(|\s*n\.|\s*$)', text, re.IGNORECASE):
        rank = m.group(1)
        cat  = m.group(2).strip()
        if rank and cat:
            results.append((rank, cat))

    if results:
        return results

    # Pattern 5 — French:  1 en Mode  or  N° 1 en Mode
    for m in re.finditer(r'(?:N°\s*)?([\d]+)\s+en\s+([^\(\n\d]+?)(?=\s*\(|\s*\d|\s*$)', text, re.IGNORECASE):
        rank = m.group(1)
        cat  = m.group(2).strip()
        if rank and cat and len(cat) > 2:
            results.append((rank, cat))

    if results:
        return results

    # Pattern 6 — Japanese:  1位 Category
    for m in re.finditer(r'([\d,]+)位\s*([^\n（\(]+?)(?=\s*（|\s*\(|\s*$|\s*[\d,]+位)', text):
        rank = m.group(1).replace(",", "")
        cat  = m.group(2).strip()
        if rank and cat:
            results.append((rank, cat))

    return results


def parse_bsr(page_content: str) -> List[Tuple[str, str]]:
    """Try multiple HTML layouts to find BSR, across all supported languages."""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(page_content, "html.parser")

    # Layout 1 — detail bullets wrapper
    wrapper = soup.find(id="detailBulletsWrapper_feature_div")
    if wrapper:
        for li in wrapper.find_all("li"):
            text = li.get_text(" ", strip=True)
            if is_bsr_label(text):
                results = extract_ranks(text)
                if results:
                    return results

    # Layout 2 — product details tables (several possible IDs)
    for tbl_id in [
        "productDetails_detailBullets_sections1",
        "productDetails_techSpec_section_1",
        "productDetails_db_sections",
    ]:
        tbl = soup.find(id=tbl_id)
        if not tbl:
            continue
        for row in tbl.find_all("tr"):
            th = row.find("th")
            td = row.find("td")
            if th and td and is_bsr_label(th.get_text()):
                results = extract_ranks(td.get_text(" ", strip=True))
                if results:
                    return results

    # Layout 3 — generic table scan (catches any table layout)
    for row in soup.find_all("tr"):
        th = row.find("th")
        td = row.find("td")
        if th and td and is_bsr_label(th.get_text()):
            results = extract_ranks(td.get_text(" ", strip=True))
            if results:
                return results

    # Layout 4 — raw text fallback (searches for any BSR label anywhere)
    body = soup.get_text(" ")
    for label in BSR_LABELS:
        idx = body.lower().find(label)
        if idx != -1:
            snippet = body[idx: idx + 500]
            results = extract_ranks(snippet)
            if results:
                return results

    return []


def get_title(page_content: str) -> str:
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(page_content, "html.parser")
    for el_id in ("productTitle", "title", "btAsinTitle"):
        el = soup.find(id=el_id)
        if el:
            return el.get_text(strip=True)
    return "N/A"


def is_captcha(page_content: str) -> bool:
    signals = [
        "robot check",
        "enter the characters you see below",
        "type the characters you see in this image",
        "sorry, we just need to make sure you're not a robot",
        "automated access to amazon",
        "api-services-support@amazon.com",
    ]
    lower = page_content.lower()
    return any(s in lower for s in signals)


# ── Scraping ──────────────────────────────────────────────────────────────────

def scrape_asin(browser, asin: str, marketplace: str) -> dict:
    domain = DOMAINS.get(marketplace.upper())
    if not domain:
        return {"status": "FAILED", "error": f"Unknown marketplace: {marketplace}", "title": "", "bsr": []}

    url    = f"{domain}/dp/{asin}"
    locale, timezone = LOCALE_MAP.get(marketplace.upper(), ("en-US", "America/New_York"))

    for attempt in range(1, MAX_RETRIES + 1):
        print(f"    Attempt {attempt}/{MAX_RETRIES} → {url}")
        context = None
        try:
            # Fresh context per attempt with the correct locale for this marketplace
            context = browser.new_context(
                viewport={"width": 1440, "height": 900},
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                locale=locale,
                timezone_id=timezone,
            )
            context.add_init_script(
                "Object.defineProperty(navigator, 'webdriver', { get: () => undefined });"
            )
            page = context.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=30000)

            # Dismiss cookie/sign-in/GDPR popups
            for selector in ["#sp-cc-accept", "input[name='accept']",
                              "#gdpr-consent-tool-wrapper button", ".a-button-close"]:
                try:
                    btn = page.locator(selector).first
                    if btn.is_visible(timeout=2000):
                        btn.click()
                        page.wait_for_timeout(500)
                except Exception:
                    pass

            # Scroll slowly to bottom
            for _ in range(8):
                page.evaluate("window.scrollBy(0, window.innerHeight * 0.8)")
                page.wait_for_timeout(random.randint(300, 600))

            # Wait for product details
            try:
                page.wait_for_selector(
                    "#detailBulletsWrapper_feature_div, "
                    "#productDetails_detailBullets_sections1, "
                    "#productDetails_techSpec_section_1",
                    timeout=8000
                )
            except PWTimeout:
                pass

            content = page.content()
            context.close()
            context = None

            if is_captcha(content):
                print(f"    CAPTCHA detected — waiting 15s before retry ...")
                time.sleep(15)
                continue

            title = get_title(content)
            bsr   = parse_bsr(content)

            if not bsr:
                print(f"    BSR not found on page — will retry")
                time.sleep(random.randint(4, 7))
                continue

            return {"status": "SUCCESS", "title": title, "bsr": bsr, "error": ""}

        except PWTimeout:
            print(f"    Page load timed out — will retry")
            time.sleep(5)
        except Exception as exc:
            print(f"    Error: {exc} — will retry")
            time.sleep(5)
        finally:
            if context:
                try:
                    context.close()
                except Exception:
                    pass

    return {"status": "FAILED", "error": "Max retries reached", "title": "", "bsr": []}


# ── I/O ───────────────────────────────────────────────────────────────────────

def read_input(filepath: str) -> List[dict]:
    wb = load_workbook(filepath)
    ws = wb.active
    products = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if len(row) < 3:
            continue
        _, asin, marketplace = row[0], row[1], row[2]
        notes = row[3] if len(row) > 3 else ""
        if not asin or not str(asin).strip():
            continue
        if not marketplace or not str(marketplace).strip():
            continue
        if notes and "example" in str(notes).lower():
            continue
        products.append({
            "asin":        str(asin).strip().upper(),
            "marketplace": str(marketplace).strip().upper(),
            "notes":       str(notes).strip() if notes else "",
        })
    return products


def write_output(results: List[dict], output_path: str) -> None:
    # ── Growisto Brand Colors (ARGB for openpyxl) ──────────────────────────
    TEAL        = "FF367588"   # headers, primary
    POWDER      = "FFB8DBD9"   # sub-headers, borders
    FLAME       = "FFE35D34"   # failed rows / alerts
    RAISIN      = "FF1D1D20"   # body text
    CULTURED    = "FFF6F6F4"   # data row bg (odd)
    WHITE       = "FFFFFFFF"   # data row bg (even) + text on dark
    POWDER_LITE = "FFE8F4F3"   # subtle alternate row

    FONT = "Poppins"

    wb = Workbook()
    ws = wb.active
    ws.title = "BSR Results"

    # Fills
    f_teal    = PatternFill("solid", fgColor=TEAL)
    f_powder  = PatternFill("solid", fgColor=POWDER)
    f_cultd   = PatternFill("solid", fgColor=CULTURED)
    f_pwdlite = PatternFill("solid", fgColor=POWDER_LITE)
    f_failed  = PatternFill("solid", fgColor="FFFDECEA")
    f_white   = PatternFill("solid", fgColor=WHITE)

    # Fonts
    fn_title  = Font(bold=True, color=WHITE,  name=FONT, size=13)
    fn_sub    = Font(bold=True, color=RAISIN, name=FONT, size=9)
    fn_header = Font(bold=True, color=WHITE,  name=FONT, size=10)
    fn_body   = Font(color=RAISIN, name=FONT, size=9)
    fn_teal   = Font(bold=True, color=TEAL,   name=FONT, size=9)
    fn_failed = Font(color=FLAME, name=FONT,  size=9)
    fn_rank   = Font(bold=True, color=TEAL,   name=FONT, size=10)

    # Borders — Powder Blue
    thin   = Side(style="thin",   color="B8DBD9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Row 1: Brand title bar ────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ts = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
    ws["A1"] = f"  Growisto BSR Report  —  {ts}"
    ws["A1"].font      = fn_title
    ws["A1"].fill      = f_teal
    ws["A1"].alignment = left
    ws.row_dimensions[1].height = 32

    # ── Row 2: Sub-header (Powder Blue) ──────────────────────────────────
    ws.merge_cells("A2:L2")
    ws["A2"] = "  Amazon Best Seller Rank Data  |  Powered by Growisto"
    ws["A2"].font      = fn_sub
    ws["A2"].fill      = f_powder
    ws["A2"].alignment = left
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers ─────────────────────────────────────────────
    headers = [
        "#", "ASIN", "Marketplace", "Product Title",
        "Main BSR Rank", "Main BSR Category",
        "Sub BSR 1 Rank", "Sub BSR 1 Category",
        "Sub BSR 2 Rank", "Sub BSR 2 Category",
        "Status", "Scraped At",
    ]
    for c, h in enumerate(headers, 1):
        cell           = ws.cell(row=3, column=c, value=h)
        cell.font      = fn_header
        cell.fill      = f_teal
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[3].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────
    for i, r in enumerate(results, 1):
        bsr   = r.get("bsr", [])
        is_ok = r["status"] == "SUCCESS"

        def safe_rank(idx):
            if len(bsr) > idx:
                v = bsr[idx][0]
                return int(v) if str(v).isdigit() else v
            return ""

        def safe_cat(idx):
            return bsr[idx][1] if len(bsr) > idx else ""

        values = [
            i,
            r["asin"],
            r["marketplace"],
            r.get("title", ""),
            safe_rank(0), safe_cat(0),
            safe_rank(1), safe_cat(1),
            safe_rank(2), safe_cat(2),
            "✓  Success" if is_ok else "✗  Failed",
            r.get("scraped_at", ""),
        ]

        row_num  = i + 3
        row_fill = (f_cultd if i % 2 != 0 else f_pwdlite) if is_ok else f_failed

        for c, val in enumerate(values, 1):
            cell        = ws.cell(row=row_num, column=c, value=val)
            cell.fill   = row_fill
            cell.border = border

            if not is_ok:
                cell.font      = fn_failed
                cell.alignment = center
            elif c == 4:                      # Product title — left aligned
                cell.font      = fn_body
                cell.alignment = left
            elif c in (5, 7, 9):              # Rank numbers — bold teal
                cell.font      = fn_rank
                cell.alignment = center
            elif c == 11:                     # Status
                cell.font      = fn_teal
                cell.alignment = center
            else:
                cell.font      = fn_body
                cell.alignment = center

        ws.row_dimensions[row_num].height = 20

    # ── Summary row ───────────────────────────────────────────────────────
    summary_row = len(results) + 4
    success_n   = sum(1 for r in results if r["status"] == "SUCCESS")
    failed_n    = len(results) - success_n
    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    ws.cell(row=summary_row, column=1,
            value=f"  Total: {len(results)}  |  Success: {success_n}  |  Failed: {failed_n}")
    ws.cell(row=summary_row, column=1).font      = Font(bold=True, color=WHITE, name=FONT, size=10)
    ws.cell(row=summary_row, column=1).fill      = f_teal
    ws.cell(row=summary_row, column=1).alignment = left
    for c in range(4, 13):
        ws.cell(row=summary_row, column=c).fill = f_teal
    ws.row_dimensions[summary_row].height = 22

    # ── Column widths ─────────────────────────────────────────────────────
    widths = [5, 16, 13, 52, 14, 32, 14, 32, 14, 32, 14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    wb.save(output_path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  default=INPUT_FILE,  help="Input Excel path")
    parser.add_argument("--output", default=None,        help="Output Excel path")
    args = parser.parse_args()

    print("=" * 60)
    print("  BSR Scraper  (Playwright / Real Browser)")
    print("=" * 60)

    if not os.path.exists(args.input):
        print(f"\nERROR: '{args.input}' not found.")
        return

    print(f"\nReading: {args.input}")
    products = read_input(args.input)

    if not products:
        print("No products found in the input file.")
        return

    print(f"Found {len(products)} product(s) to scrape.")
    print("Opening browser...\n")
    print("-" * 60)

    results = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled",
            ]
        )

        for idx, p in enumerate(products, 1):
            print(f"\n[{idx}/{len(products)}]  ASIN: {p['asin']}  |  Marketplace: {p['marketplace']}")

            result               = scrape_asin(browser, p["asin"], p["marketplace"])
            result["asin"]       = p["asin"]
            result["marketplace"]= p["marketplace"]
            result["scraped_at"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            if result["status"] == "SUCCESS":
                bsr = result.get("bsr", [])
                print(f"    OK  {result.get('title', '')[:65]}")
                if bsr:
                    print(f"    BSR: #{bsr[0][0]}  in  {bsr[0][1]}")
                    for sub in bsr[1:]:
                        print(f"         #{sub[0]}  in  {sub[1]}")
            else:
                print(f"    FAILED — {result.get('error', '')}")

            results.append(result)

            if idx < len(products):
                wait = random.uniform(3, 6)
                print(f"    Waiting {wait:.1f}s ...")
                time.sleep(wait)

        browser.close()


    timestamp   = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = args.output or f"BSR_Output_{timestamp}.xlsx"
    write_output(results, output_file)

    success = sum(1 for r in results if r["status"] == "SUCCESS")
    failed  = len(results) - success

    print("\n" + "=" * 60)
    print(f"  Done.  Success: {success}  |  Failed: {failed}")
    print(f"  Output saved: {output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
